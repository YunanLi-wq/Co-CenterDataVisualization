from flask import Flask, render_template, send_from_directory, jsonify, Response, request, redirect, url_for
import os
from pymongo import MongoClient
from bson import ObjectId
import json
import re
import unicodedata
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import io
from datetime import datetime

app = Flask(__name__)

# MongoDB Connection
try:
    client = MongoClient('mongodb://127.0.0.1:27017/')
    db = client['local']
    collection = db['rolNLDraft']
    researcher_collection = db['Researcher']  # Collection for researcher profiles linked to Factor
    manage_collection = db['Manager']  # Collection for user authentication
    pending_collection = db['PendingItems']  # Collection for pending items awaiting approval
    higher_manager_collection = db['HigherManager']  # Collection for higher manager authentication
    description_collection = db['description']  # Collection for module descriptions
    compass_collection = db['Compass']  # Collection for Compass structure (Quadrant, Segment, Factor)
    mongo_connected = True
except Exception as e:
    print(f"MongoDB connection failed: {e}")
    mongo_connected = False

@app.route('/')
def index():
    """Serve the main HTML page"""
    return render_template('index.html')

@app.route('/d3viz')
def d3viz():
    """Serve the Food Co-Centre Sustainability Compass visualization page"""
    # The up-to-date Compass UI lives in d3viz2.html
    return render_template('d3viz2.html')

@app.route('/intro')
def intro():
    """Serve the Compass introduction page."""
    return render_template('intro.html')

@app.route('/d3viz2')
def d3viz2():
    """Serve the Food Co-Centre Sustainability Compass visualization page (with per-wedge count bubbles)"""
    return render_template('d3viz2.html')

@app.route('/api/dataset/module-counts')
def get_module_counts():
    """Return record counts grouped by Quadrant/Segment/Factor for visualization overlays."""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})

        def group_counts(field_name: str):
            pipeline = [
                {'$match': {field_name: {'$exists': True, '$ne': None, '$ne': ''}}},
                {'$group': {'_id': f'${field_name}', 'count': {'$sum': 1}}},
                {'$sort': {'count': -1}}
            ]
            rows = list(collection.aggregate(pipeline))
            result = {}
            for r in rows:
                key = r.get('_id')
                if key is None:
                    continue
                key_str = str(key).strip()
                if not key_str:
                    continue
                result[key_str] = int(r.get('count', 0))
            return result

        return jsonify({
            'success': True,
            'counts': {
                'quadrant': group_counts('Quadrant'),
                'segment': group_counts('Segment'),
                'factor': group_counts('Factor')
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/researchers/factor-counts')
def get_researcher_factor_counts():
    """Return researcher counts grouped by Factor (for d3viz2 outer bubbles)."""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})

        # Optional filters (Researcher collection schema)
        # - Platform Alignment may contain multiple comma-separated values, e.g. "P1.1, P3"
        platform_alignment = request.args.get('platform_alignment', '').strip()
        institution = request.args.get('institution', '').strip()
        funder_category = request.args.get('funder_category', '').strip()

        def exact_value_regex(value: str):
            """Case-insensitive exact match, tolerant of leading/trailing whitespace in stored values."""
            v = (value or '').strip()
            if not v:
                return None
            return {"$regex": rf'^\s*{re.escape(v)}\s*$', "$options": "i"}

        def platform_alignment_regex(value: str):
            """Match a platform value inside a comma-separated 'Platform Alignment' field.
            Examples:
              value='P1' matches 'P1' and also 'P1.1' inside 'P1.1, P3'
              value='P3' matches '..., P3' inside 'P1.1, P3'
            """
            v = (value or '').strip()
            if not v:
                return None
            # (^|,)\s*<v>(\b|[.\d]) ensures token-boundary-ish matching and supports 'P1.1'
            return {"$regex": rf'(^|,)\s*{re.escape(v)}(\b|[.\d])', "$options": "i"}

        # Build base filter clauses (so we can safely combine multiple $or constraints via $and)
        base_clauses = []
        if platform_alignment:
            rx = platform_alignment_regex(platform_alignment)
            if rx:
                base_clauses.append({'Platform Alignment': rx})
        if institution:
            rx = exact_value_regex(institution)
            if rx:
                base_clauses.append({'$or': [
                    {'Institution': rx},
                    {'Institution ': rx},
                    {'University': rx},
                    {'University ': rx},
                ]})
        if funder_category:
            rx = exact_value_regex(funder_category)
            if rx:
                # Career Stage is stored in "Column1" (backward compatible with old "Funder Category")
                base_clauses.append({'$or': [
                    {'Column1': rx},
                    {'Column1 ': rx},
                    {'Funder Category': rx},
                    {'Funder Category ': rx},
                ]})

        match_conditions = {'$and': [{'Factor': {'$exists': True, '$ne': None, '$ne': ''}}] + base_clauses} if base_clauses else {'Factor': {'$exists': True, '$ne': None, '$ne': ''}}

        # all_factor_count is computed in wide-schema section (supports AllFacor/AllFactor + safe $or merging)
        all_factor_count = 0

        # Researchers can be stored in two schemas:
        # 1) Narrow: a single 'Factor' field containing the factor name (string).
        # 2) Wide (one-hot): each Factor is a column; membership indicated by value 1 (or "1"/true).
        #
        # We'll detect narrow schema quickly; if not found, fall back to wide schema counting.

        has_narrow_factor = researcher_collection.find_one(
            {'Factor': {'$exists': True, '$ne': None, '$ne': ''}},
            {'_id': 1}
        ) is not None

        if has_narrow_factor:
            pipeline = [
                {'$match': match_conditions},
                {'$group': {'_id': '$Factor', 'count': {'$sum': 1}}},
                {'$sort': {'count': -1}}
            ]
            rows = list(researcher_collection.aggregate(pipeline))
            counts = {}
            for r in rows:
                key = r.get('_id')
                if key is None:
                    continue
                k = str(key).strip()
                if not k:
                    continue
                counts[k] = int(r.get('count', 0))
            # Center AllFactor count for narrow schema (respects filters)
            try:
                membership_values_any = [1, 1.0, '1', '1.0', True]
                all_q = {'$and': base_clauses + [{'AllFactor': {'$in': membership_values_any}}]} if base_clauses else {'AllFactor': {'$in': membership_values_any}}
                all_factor_count = int(researcher_collection.count_documents(all_q))
            except Exception:
                all_factor_count = 0
            return jsonify({'success': True, 'counts': counts, 'key_map': {}, 'all_factor_count': all_factor_count})

        # Wide schema: compute counts by counting documents where the factor-column == 1
        def clean_list(values):
            out = []
            seen = set()
            for v in values or []:
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue
                key = s.lower()
                if key in seen:
                    continue
                seen.add(key)
                out.append(s)
            out.sort()
            return out

        def _apply_known_factor_aliases(s: str) -> str:
            # Known aliases / corrections (apply before normalization)
            s2 = s
            s2 = s2.replace('genetic biodiversity', 'genetic diversity')
            s2 = s2.replace('phosphorus', 'phosphorous')
            # Common misspelling in datasets
            s2 = s2.replace('contaminents', 'contaminants')
            return s2

        def normalize_factor_keep_spaces(s: str) -> str:
            """Step 1 (user rule): remove symbols but DO NOT remove spaces.

            - Unicode-normalize (NFKC)
            - Lowercase
            - Delete punctuation/symbols; keep letters/numbers and whitespace
            - Collapse whitespace to single spaces
            """
            if s is None:
                return ''
            s2 = unicodedata.normalize('NFKC', str(s)).strip().lower()
            s2 = _apply_known_factor_aliases(s2)
            out = []
            for ch in s2:
                if ch.isspace():
                    out.append(' ')
                    continue
                cat = unicodedata.category(ch)
                if cat and cat[0] in ('L', 'N'):
                    out.append(ch)
                # else: delete symbol/punctuation entirely
            return ' '.join(''.join(out).split())

        def normalize_factor_compact(s: str) -> str:
            """Step 2 (fallback): remove symbols AND spaces (keep only letters/numbers)."""
            keep_spaces = normalize_factor_keep_spaces(s)
            return keep_spaces.replace(' ', '')

        def infer_factor_columns_from_researchers():
            """Best-effort inference of one-hot Factor columns from a sample Researcher doc."""
            sample = researcher_collection.find_one({}, {'_id': 0}) or {}
            exclude = {
                'Name',
                'Institution',
                'Funder Category',
                'Platform Alignment',
                'Email (as per SESAME)',
                'Co-Centre Role',
                'Brief description of intervention(s) they are working on',
                'Brief Description',
                'Factor',
                # All-factors marker (should not be treated as a real Factor column)
                'AllFactor'
            }
            candidates = []
            for k, v in sample.items():
                if not k or k.startswith('_') or k in exclude:
                    continue
                # Heuristic: factor columns are typically 0/1-ish values
                if v in (0, 1, '0', '1', True, False, None, ''):
                    candidates.append(str(k).strip())
            return clean_list(candidates)

        # Prefer Compass collection as authoritative list of *display* factor names
        display_factors = []
        try:
            display_factors = clean_list(compass_collection.distinct('Factor'))
        except Exception:
            display_factors = []

        # Infer one-hot factor columns directly from Researcher collection
        factor_columns = infer_factor_columns_from_researchers()
        # If Compass is empty, fall back to using columns as display labels too
        if not display_factors:
            display_factors = factor_columns

        counts = {}
        key_map = {}  # display factor -> actual field/column name used for counting
        base_query = {'$and': base_clauses} if base_clauses else {}

        membership_values = [1, 1.0, '1', '1.0', True]

        # Cache for field existence checks (avoids repeated scans)
        _exists_cache = {}

        def field_exists(field_name: str) -> bool:
            key = str(field_name or '').strip()
            if not key:
                return False
            if key in _exists_cache:
                return _exists_cache[key]
            try:
                exists = researcher_collection.find_one({key: {'$exists': True}}, {'_id': 1}) is not None
            except Exception:
                exists = False
            _exists_cache[key] = exists
            return exists

        # All-factors marker key
        all_factor_keys = ['AllFactor']

        def all_factor_or_clause():
            return [{k: {'$in': membership_values}} for k in all_factor_keys if field_exists(k)]

        # Count of AllFactor for center bubble (respects filters)
        all_factor_count = 0
        try:
            base_for_all = dict(base_query) if base_query else {}
            all_or = all_factor_or_clause()
            if all_or:
                all_q = {'$and': [base_for_all, {'$or': all_or}]} if base_for_all else {'$or': all_or}
                all_factor_count = int(researcher_collection.count_documents(all_q))
        except Exception:
            all_factor_count = 0

        def factor_key_candidates(display_name: str):
            """Generate likely one-hot column names for a given Factor display label."""
            base = str(display_name or '').strip()
            if not base:
                return []
            cands = []

            def add(x):
                x = str(x or '').strip()
                if x and x not in cands:
                    cands.append(x)

            def strip_symbols_keep_spaces_preserve_case(s: str) -> str:
                """Delete all symbols/punctuation except spaces, and collapse whitespace.

                This is used to match display labels like:
                  'Foodborne microbiological， physical ...'  ->  'Foodborne microbiological physical ...'
                so they can match one-hot column names that omit punctuation.
                """
                s2 = unicodedata.normalize('NFKC', str(s or '')).strip()
                out = []
                for ch in s2:
                    if ch.isspace():
                        out.append(' ')
                        continue
                    cat = unicodedata.category(ch)  # e.g. 'Ll', 'Po', 'Sc'
                    if cat and cat[0] in ('L', 'N'):  # letters / numbers
                        out.append(ch)
                return ' '.join(''.join(out).split())

            add(base)
            # Apostrophe variants
            add(base.replace('’', "'"))
            add(base.replace("'", "’"))
            # Comma variants (e.g. "a, b" vs "a b")
            add(re.sub(r',\s*', ' ', base))
            add(re.sub(r',\s*', '', base))
            # Normalize whitespace
            add(re.sub(r'\s+', ' ', base))
            # Remove most punctuation entirely (keeps letters/numbers/spaces)
            add(re.sub(r'[^A-Za-z0-9 ]+', ' ', base))
            add(re.sub(r'\s+', ' ', re.sub(r'[^A-Za-z0-9 ]+', ' ', base)).strip())
            # Unicode punctuation-insensitive candidate (handles Chinese commas etc.)
            add(' '.join(unicodedata.normalize('NFKC', base).split()))
            # Step 1 (user rule): delete symbols, keep spaces
            add(strip_symbols_keep_spaces_preserve_case(base))
            # Step 2 (fallback): delete symbols AND spaces
            add(strip_symbols_keep_spaces_preserve_case(base).replace(' ', ''))

            return cands

        # Build normalized lookup of available columns so punctuation/spacing differences still match
        col_by_keep_spaces = {}
        col_by_compact = {}
        for col in factor_columns:
            k1 = normalize_factor_keep_spaces(col)
            if k1:
                # Prefer the first seen; most datasets won't have true collisions
                col_by_keep_spaces.setdefault(k1, col)
            k2 = normalize_factor_compact(col)
            if k2:
                col_by_compact.setdefault(k2, col)

        for f in display_factors:
            # Find the actual column to count on (wide schema)
            col = None
            if f in factor_columns:
                col = f
            else:
                # 1) Remove symbols but KEEP spaces (user rule)
                col = col_by_keep_spaces.get(normalize_factor_keep_spaces(f))
                if not col:
                    # 2) Remove symbols AND spaces (user rule)
                    col = col_by_compact.get(normalize_factor_compact(f))

            # Fallback: the column may exist but wasn't present in our sample document.
            # Try a few likely name variants and pick the first that exists.
            if not col:
                for cand in factor_key_candidates(f):
                    if field_exists(cand):
                        col = cand
                        break

            if not col:
                counts[f] = 0
                continue

            # Factor bubble counts should EXCLUDE AllFactor researchers (they are shown in the center bubble only).
            # Combine base filters with:
            # - this Factor membership
            # - AllFactor != 1
            clauses = []
            if base_query:
                clauses.append(dict(base_query))
            clauses.append({col: {'$in': membership_values}})
            clauses.append({'AllFactor': {'$nin': membership_values}})
            q = {'$and': clauses} if len(clauses) > 1 else (clauses[0] if clauses else {})
            try:
                c = int(researcher_collection.count_documents(q))
                # Only return factors that have at least 1 researcher under current filters
                if c > 0:
                    counts[f] = c
                    key_map[f] = col
            except Exception:
                pass

        return jsonify({'success': True, 'counts': counts, 'key_map': key_map, 'all_factor_count': all_factor_count})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/researchers/by-factor')
def get_researchers_by_factor():
    """Return researchers for a given Factor (Name, Institution, Funder Category, Platform Alignment)."""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})

        factor = request.args.get('factor', '').strip()
        if not factor:
            return jsonify({'success': False, 'error': 'No factor parameter provided'})
        # In wide schema, the factor column name may differ from the display name
        factor_key = request.args.get('factor_key', '').strip()

        # Optional filters (Researcher collection schema)
        platform_alignment = request.args.get('platform_alignment', '').strip()
        institution = request.args.get('institution', '').strip()
        funder_category = request.args.get('funder_category', '').strip()

        def exact_value_regex(value: str):
            v = (value or '').strip()
            if not v:
                return None
            return {"$regex": rf'^\s*{re.escape(v)}\s*$', "$options": "i"}

        def platform_alignment_regex(value: str):
            v = (value or '').strip()
            if not v:
                return None
            return {"$regex": rf'(^|,)\s*{re.escape(v)}(\b|[.\d])', "$options": "i"}

        # Detect schema (narrow Factor field vs wide one-hot factor columns)
        has_narrow_factor = researcher_collection.find_one(
            {'Factor': {'$exists': True, '$ne': None, '$ne': ''}},
            {'_id': 1}
        ) is not None

        if has_narrow_factor:
            query = {'Factor': {"$regex": f'^{re.escape(factor)}$', "$options": "i"}}
        else:
            # Wide schema: factor name is a field/column; membership indicated by value 1
            key = factor_key or factor
            # If the provided key doesn't exist (punctuation/space variations), try to resolve it.
            def field_exists(field_name: str) -> bool:
                k = str(field_name or '').strip()
                if not k:
                    return False
                try:
                    return researcher_collection.find_one({k: {'$exists': True}}, {'_id': 1}) is not None
                except Exception:
                    return False

            def factor_key_candidates(display_name: str):
                base = str(display_name or '').strip()
                if not base:
                    return []
                cands = []
                def add(x):
                    x = str(x or '').strip()
                    if x and x not in cands:
                        cands.append(x)
                def strip_symbols_keep_spaces_preserve_case(s: str) -> str:
                    s2 = unicodedata.normalize('NFKC', str(s or '')).strip()
                    out = []
                    for ch in s2:
                        if ch.isspace():
                            out.append(' ')
                            continue
                        cat = unicodedata.category(ch)
                        if cat and cat[0] in ('L', 'N'):
                            out.append(ch)
                    return ' '.join(''.join(out).split())
                add(base)
                add(base.replace('’', "'"))
                add(base.replace("'", "’"))
                add(re.sub(r',\s*', ' ', base))
                add(re.sub(r'\s+', ' ', base))
                add(re.sub(r'\s+', ' ', re.sub(r'[^A-Za-z0-9 ]+', ' ', base)).strip())
                # Unicode punctuation-insensitive candidate (handles Chinese commas etc.)
                try:
                    add(' '.join(unicodedata.normalize('NFKC', base).split()))
                except Exception:
                    pass
                # Normalized form (punctuation stripped)
                try:
                    add(' '.join(''.join((ch if (unicodedata.category(ch)[:1] in ('L','N')) else ' ') for ch in unicodedata.normalize('NFKC', base).lower()).split()))
                except Exception:
                    pass
                # Preserve original casing while deleting punctuation/symbols (user rule)
                try:
                    add(strip_symbols_keep_spaces_preserve_case(base))
                except Exception:
                    pass
                # Fallback: delete punctuation/symbols AND spaces
                try:
                    add(strip_symbols_keep_spaces_preserve_case(base).replace(' ', ''))
                except Exception:
                    pass
                return cands

            if not field_exists(key):
                for cand in factor_key_candidates(key):
                    if field_exists(cand):
                        key = cand
                        break
            # Factor list should EXCLUDE AllFactor researchers (they are shown via center bubble only).
            membership_values = [1, 1.0, '1', '1.0', True]
            query = {
                '$and': [
                    {key: {'$in': membership_values}},
                    {'AllFactor': {'$nin': membership_values}}
                ]
            }

        # Apply optional filters as additional AND clauses (avoid overwriting $or from other logic)
        and_clauses = [query] if query else []
        if platform_alignment:
            rx = platform_alignment_regex(platform_alignment)
            if rx:
                and_clauses.append({'Platform Alignment': rx})
        if institution:
            rx = exact_value_regex(institution)
            if rx:
                and_clauses.append({'$or': [
                    {'Institution': rx},
                    {'Institution ': rx},
                    {'University': rx},
                    {'University ': rx},
                ]})
        if funder_category:
            rx = exact_value_regex(funder_category)
            if rx:
                and_clauses.append({'$or': [
                    {'Column1': rx},
                    {'Column1 ': rx},
                    {'Funder Category': rx},
                    {'Funder Category ': rx},
                ]})
        query = {'$and': and_clauses} if len(and_clauses) > 1 else (and_clauses[0] if and_clauses else {})

        projection = {
            '_id': 0,
            'Name': 1,
            'Institution': 1,
            'Funder Category': 1,
            'Funder Category ': 1,
            'Column1': 1,
            'Column1 ': 1,
            'Platform Alignment': 1,
            'Email (as per SESAME)': 1,
            'Co-Centre Role': 1,
            'Brief description of intervention(s) they are working on': 1,
            'Factor': 1
        }
        docs = list(researcher_collection.find(query, projection))

        # Normalize output fields
        def norm(doc):
            career_stage = doc.get('Column1')
            if career_stage is None:
                career_stage = doc.get('Column1 ')
            if career_stage is None:
                career_stage = doc.get('Funder Category')
            if career_stage is None:
                career_stage = doc.get('Funder Category ')
            return {
                'Name': doc.get('Name', '') or '',
                'Institution': doc.get('Institution', '') or '',
                # Keep response key stable for frontend; underlying field is now "Column1"
                'Funder Category': career_stage or '',
                'Email (as per SESAME)': doc.get('Email (as per SESAME)', '') or '',
                'Platform Alignment': doc.get('Platform Alignment', '') or '',
                'Co-Centre Role': doc.get('Co-Centre Role', '') or '',
                # Simplified field name for UI
                'Brief Description': doc.get('Brief description of intervention(s) they are working on', '') or ''
            }

        researchers = [norm(d) for d in docs]
        return jsonify({'success': True, 'factor': factor, 'count': len(researchers), 'researchers': researchers})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/researchers/allfactor')
def get_researchers_allfactor():
    """Return researchers where AllFactor == 1 (respects optional filters)."""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})

        platform_alignment = request.args.get('platform_alignment', '').strip()
        institution = request.args.get('institution', '').strip()
        funder_category = request.args.get('funder_category', '').strip()

        def exact_value_regex(value: str):
            v = (value or '').strip()
            if not v:
                return None
            return {"$regex": rf'^\s*{re.escape(v)}\s*$', "$options": "i"}

        def platform_alignment_regex(value: str):
            v = (value or '').strip()
            if not v:
                return None
            return {"$regex": rf'(^|,)\s*{re.escape(v)}(\b|[.\d])', "$options": "i"}

        membership_values = [1, 1.0, '1', '1.0', True]
        and_clauses = [{'AllFactor': {'$in': membership_values}}]

        if platform_alignment:
            rx = platform_alignment_regex(platform_alignment)
            if rx:
                and_clauses.append({'Platform Alignment': rx})
        if institution:
            rx = exact_value_regex(institution)
            if rx:
                # Backward compatible: some datasets may still use University or trailing-space variants
                and_clauses.append({'$or': [
                    {'Institution': rx},
                    {'Institution ': rx},
                    {'University': rx},
                    {'University ': rx},
                ]})
        if funder_category:
            rx = exact_value_regex(funder_category)
            if rx:
                and_clauses.append({'$or': [
                    {'Column1': rx},
                    {'Column1 ': rx},
                    {'Funder Category': rx},
                    {'Funder Category ': rx},
                ]})
        query = {'$and': and_clauses} if len(and_clauses) > 1 else and_clauses[0]

        projection = {
            '_id': 0,
            'Name': 1,
            'Institution': 1,
            'Funder Category': 1,
            'Funder Category ': 1,
            'Column1': 1,
            'Column1 ': 1,
            'Email (as per SESAME)': 1,
            'Platform Alignment': 1,
            'Co-Centre Role': 1,
            'Brief description of intervention(s) they are working on': 1,
        }
        docs = list(researcher_collection.find(query, projection))

        def norm(doc):
            career_stage = doc.get('Column1')
            if career_stage is None:
                career_stage = doc.get('Column1 ')
            if career_stage is None:
                career_stage = doc.get('Funder Category')
            if career_stage is None:
                career_stage = doc.get('Funder Category ')
            return {
                'Name': doc.get('Name', '') or '',
                'Institution': doc.get('Institution', '') or '',
                'Funder Category': career_stage or '',
                'Email (as per SESAME)': doc.get('Email (as per SESAME)', '') or '',
                'Platform Alignment': doc.get('Platform Alignment', '') or '',
                'Co-Centre Role': doc.get('Co-Centre Role', '') or '',
                'Brief Description': doc.get('Brief description of intervention(s) they are working on', '') or ''
            }

        researchers = [norm(d) for d in docs]
        return jsonify({'success': True, 'count': len(researchers), 'researchers': researchers})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/researchers/all')
def get_researchers_all():
    """Return all researchers (respects optional filters)."""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})

        platform_alignment = request.args.get('platform_alignment', '').strip()
        institution = request.args.get('institution', '').strip()
        funder_category = request.args.get('funder_category', '').strip()

        def exact_value_regex(value: str):
            v = (value or '').strip()
            if not v:
                return None
            return {"$regex": rf'^\s*{re.escape(v)}\s*$', "$options": "i"}

        def platform_alignment_regex(value: str):
            v = (value or '').strip()
            if not v:
                return None
            return {"$regex": rf'(^|,)\s*{re.escape(v)}(\b|[.\d])', "$options": "i"}

        and_clauses = []
        if platform_alignment:
            rx = platform_alignment_regex(platform_alignment)
            if rx:
                and_clauses.append({'Platform Alignment': rx})
        if institution:
            rx = exact_value_regex(institution)
            if rx:
                and_clauses.append({'$or': [
                    {'Institution': rx},
                    {'Institution ': rx},
                    {'University': rx},
                    {'University ': rx},
                ]})
        if funder_category:
            rx = exact_value_regex(funder_category)
            if rx:
                and_clauses.append({'$or': [
                    {'Column1': rx},
                    {'Column1 ': rx},
                    {'Funder Category': rx},
                    {'Funder Category ': rx},
                ]})
        query = {'$and': and_clauses} if and_clauses else {}

        projection = {
            '_id': 0,
            'Name': 1,
            'Institution': 1,
            'Funder Category': 1,
            'Funder Category ': 1,
            'Column1': 1,
            'Column1 ': 1,
            'Email (as per SESAME)': 1,
            'Platform Alignment': 1,
            'Co-Centre Role': 1,
            'Brief description of intervention(s) they are working on': 1,
        }
        docs = list(researcher_collection.find(query, projection))

        def norm(doc):
            career_stage = doc.get('Column1')
            if career_stage is None:
                career_stage = doc.get('Column1 ')
            if career_stage is None:
                career_stage = doc.get('Funder Category')
            if career_stage is None:
                career_stage = doc.get('Funder Category ')
            return {
                'Name': doc.get('Name', '') or '',
                'Institution': doc.get('Institution', '') or '',
                'Funder Category': career_stage or '',
                'Email (as per SESAME)': doc.get('Email (as per SESAME)', '') or '',
                'Platform Alignment': doc.get('Platform Alignment', '') or '',
                'Co-Centre Role': doc.get('Co-Centre Role', '') or '',
                'Brief Description': doc.get('Brief description of intervention(s) they are working on', '') or ''
            }

        researchers = [norm(d) for d in docs]
        return jsonify({'success': True, 'count': len(researchers), 'researchers': researchers})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/researchers/platform-summary')
def get_researcher_platform_summary():
    """Return total researchers + platform counts (P1..P5).

    Notes:
    - Applies optional filters: platform_alignment / institution / funder_category.
    - If scope=factor, requires factor_key and excludes AllFactor=1 (so it matches factor bubbles).
    - If scope=allfactor, counts only AllFactor=1.
    - Platform counts are NOT mutually exclusive (a person can be counted in multiple platforms).
    """
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})

        scope = request.args.get('scope', 'all').strip().lower()  # all | factor | allfactor
        factor_key = request.args.get('factor_key', '').strip()

        platform_alignment = request.args.get('platform_alignment', '').strip()
        institution = request.args.get('institution', '').strip()
        funder_category = request.args.get('funder_category', '').strip()

        def exact_value_regex(value: str):
            v = (value or '').strip()
            if not v:
                return None
            return {"$regex": rf'^\s*{re.escape(v)}\s*$', "$options": "i"}

        def platform_alignment_regex(value: str):
            v = (value or '').strip()
            if not v:
                return None
            return {"$regex": rf'(^|,)\s*{re.escape(v)}(\b|[.\d])', "$options": "i"}

        membership_values = [1, 1.0, '1', '1.0', True]
        and_clauses = []

        # Scope filter
        if scope == 'allfactor':
            and_clauses.append({'AllFactor': {'$in': membership_values}})
        elif scope == 'factor':
            if not factor_key:
                return jsonify({'success': False, 'error': 'factor_key is required when scope=factor'})
            and_clauses.append({factor_key: {'$in': membership_values}})
            # Factor bubbles exclude AllFactor researchers
            and_clauses.append({'AllFactor': {'$nin': membership_values}})
        else:
            # all researchers (no scope clause)
            pass

        # Optional filters
        if platform_alignment:
            rx = platform_alignment_regex(platform_alignment)
            if rx:
                and_clauses.append({'Platform Alignment': rx})
        if institution:
            rx = exact_value_regex(institution)
            if rx:
                and_clauses.append({'$or': [
                    {'Institution': rx},
                    {'Institution ': rx},
                    {'University': rx},
                    {'University ': rx},
                ]})
        if funder_category:
            rx = exact_value_regex(funder_category)
            if rx:
                and_clauses.append({'$or': [
                    {'Column1': rx},
                    {'Column1 ': rx},
                    {'Funder Category': rx},
                    {'Funder Category ': rx},
                ]})

        query = {'$and': and_clauses} if and_clauses else {}

        # Only need Platform Alignment for counting
        docs = list(researcher_collection.find(query, {'_id': 0, 'Platform Alignment': 1}))

        def platforms_for_doc(doc):
            raw = doc.get('Platform Alignment')
            if raw is None:
                return set()
            s = str(raw).strip()
            if not s:
                return set()
            parts = [p.strip() for p in s.split(',') if p.strip()]
            out = set()
            for p in parts:
                m = re.match(r'^(P\d+)', p, flags=re.IGNORECASE)
                if m:
                    out.add(m.group(1).upper())
            return out

        total = len(docs)
        platform_counts = {f'P{i}': 0 for i in range(1, 6)}
        for d in docs:
            for p in platforms_for_doc(d):
                if p in platform_counts:
                    platform_counts[p] += 1

        return jsonify({
            'success': True,
            'total': total,
            'platform_counts': platform_counts
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/researchers/unique-values')
def get_researcher_unique_values():
    """Return unique Platform Alignment / Institution / Funder Category values for filter dropdowns in d3viz2."""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})

        def clean(values):
            out = []
            seen = set()
            for v in values:
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue
                key = s.lower()
                if key in seen:
                    continue
                seen.add(key)
                out.append(s)
            out.sort()
            return out

        # Only include Platform Alignment / Institution / Career Stage values for researchers
        # that actually have at least one *real Factor* membership (one-hot column == 1).
        #
        # Important: the Researcher dataset may contain other 0/1 fields like
        # "Completed survey to some extent" or "Contacted?" which must NOT be treated as Factor membership.
        # We therefore gate membership by checking keys against the Compass Factor list.
        membership_values = {1, 1.0, '1', '1.0', True}
        exclude_fields = {
            'Name',
            'Institution',
            'Institution ',
            'University',
            'University ',
            'Funder Category',
            'Funder Category ',
            'Column1',
            'Column1 ',
            'Platform Alignment',
            'Email (as per SESAME)',
            'Co-Centre Role',
            'Brief description of intervention(s) they are working on',
            'Brief Description',
            'Factor'
        }

        # Build a normalized set of known Factor names from Compass collection
        def norm_factor_key(s: str) -> str:
            # Lowercase, remove punctuation/symbols, collapse spaces
            s2 = re.sub(r'[^0-9a-zA-Z\\s]+', ' ', str(s or '')).lower()
            return ' '.join(s2.split())

        try:
            compass_factors = [f for f in compass_collection.distinct('Factor') if f]
        except Exception:
            compass_factors = []
        factor_norm_set = {norm_factor_key(f) for f in compass_factors if norm_factor_key(f)}

        platform_tokens = []
        p1_sub_tokens = []
        institutions_with_factors = []
        funder_categories_with_factors = []
        seen_inst = set()
        seen_fc = set()

        def has_factor_membership(doc: dict) -> bool:
            # If we have a Compass factor list, only count membership on those keys
            if factor_norm_set:
                for k, v in (doc or {}).items():
                    if not k or k in exclude_fields or str(k).startswith('_'):
                        continue
                    if v not in membership_values:
                        continue
                    if norm_factor_key(k) in factor_norm_set:
                        return True
                return False

            # Fallback (if Compass factors not available): previous heuristic, plus exclude common non-factor fields
            non_factor_like = {
                'Actions',
                'Contacted?',
                'Completed survey to some extent'
            }
            for k, v in (doc or {}).items():
                if not k or k in exclude_fields or k in non_factor_like or str(k).startswith('_'):
                    continue
                if v in membership_values:
                    return True
            return False

        for doc in researcher_collection.find({}, {'_id': 0}):
            if not has_factor_membership(doc):
                continue

            # Platform Alignment: may store multiple values like "P1.1, P3"
            pa = doc.get('Platform Alignment')
            if pa is not None:
                s = str(pa).strip()
                if s:
                    parts = [p.strip() for p in s.split(',') if p.strip()]
                    for p in parts:
                        # Top-level platform prefix, e.g. P1 from P1.1
                        m = re.match(r'^(P\d+)', p, flags=re.IGNORECASE)
                        if m:
                            platform_tokens.append(m.group(1).upper())
                        else:
                            platform_tokens.append(p)

                        # Collect P1 sub-platforms (e.g. P1.1, P1.2 ...)
                        msub = re.match(r'^(P1)\.(\d+)$', p, flags=re.IGNORECASE)
                        if msub:
                            p1_sub_tokens.append(f"P1.{int(msub.group(2))}")

            # Institution: support legacy fields + trailing-space variants
            for field in ('Institution', 'Institution ', 'University', 'University '):
                v = doc.get(field)
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue
                key = s.lower()
                if key in seen_inst:
                    continue
                seen_inst.add(key)
                institutions_with_factors.append(s)

            # Career Stage (stored in "Column1"; fallback to old "Funder Category")
            v = doc.get('Column1')
            if v is None:
                v = doc.get('Column1 ')
            if v is None:
                v = doc.get('Funder Category')
            if v is None:
                v = doc.get('Funder Category ')
            if v is not None:
                s = str(v).strip()
                if s:
                    key = s.lower()
                    if key not in seen_fc:
                        seen_fc.add(key)
                        funder_categories_with_factors.append(s)

        platform_alignments = clean(platform_tokens)

        # Sort P1 sub-options numerically by suffix
        def sort_p1_sub(values):
            uniq = clean(values)

            def keyfn(x):
                m = re.match(r'^P1\.(\d+)$', x.strip(), flags=re.IGNORECASE)
                return int(m.group(1)) if m else 10**9

            return sorted(uniq, key=keyfn)

        platform_p1_sub_alignments = sort_p1_sub(p1_sub_tokens)
        institutions = clean(institutions_with_factors)
        funder_categories = clean(funder_categories_with_factors)

        return jsonify({
            'success': True,
            'platform_alignments': platform_alignments,
            'platform_p1_sub_alignments': platform_p1_sub_alignments,
            'institutions': institutions,
            'funder_categories': funder_categories
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/table/<path:module_name>')
def table(module_name):
    """Serve the data table page for a specific module"""
    # Decode the module name in case it was URL encoded
    import urllib.parse
    decoded_module_name = urllib.parse.unquote(module_name)
    return render_template('table.html', module_name=decoded_module_name)

@app.route('/login')
def login():
    """Serve the login page"""
    return render_template('login.html')

@app.route('/add')
def add():
    """Serve the add item page"""
    return render_template('add.html')

@app.route('/update')
def update():
    """Serve the update/review page"""
    return render_template('update.html')

@app.route('/higher-manager-login')
def higher_manager_login():
    """Serve the higher manager login page"""
    return render_template('higher_manager_login.html')

@app.route('/static/<path:filename>')
def static_files(filename):
    """Serve static files like CSS"""
    return send_from_directory('static', filename)

@app.route('/figure/<path:filename>')
def figure_files(filename):
    """Serve figure images used by intro page"""
    return send_from_directory('figure', filename)

@app.route('/api/dataset/status')
def dataset_status():
    """Check database connection status"""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})
        
        count = collection.count_documents({})
        return jsonify({'success': True, 'count': count})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/dataset/load')
def load_dataset():
    """Load dataset"""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})
        
        documents = list(collection.find().limit(10))  # Limit to 10 documents
        count = collection.count_documents({})
        
        # Convert ObjectId to string
        for doc in documents:
            if '_id' in doc:
                doc['_id'] = str(doc['_id'])
        
        return jsonify({'success': True, 'documents': documents, 'count': count})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/dataset/view')
def view_collection():
    """View collection data"""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})
        
        documents = list(collection.find().limit(20))  # Limit to 20 documents
        
        # Convert ObjectId to string
        for doc in documents:
            if '_id' in doc:
                doc['_id'] = str(doc['_id'])
        
        return jsonify({'success': True, 'documents': documents})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/dataset/excel-preview')
def excel_preview():
    """Get Excel preview data"""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})
        
        documents = list(collection.find().limit(50))  # Limit for preview
        
        if not documents:
            # If no documents, show Food Co-Centre Sustainability Compass structure as example
            food_compass_data = [
                {
                    "societal_goal": "Safe & healthy diets for all",
                    "name": "Safe & healthy diets for all",
                    "type": "Quadrant",
                    "level": 1,
                    "children": [
                        {"societal_goal": "Safe & healthy diets for all", "name": "Food intake", "type": "Segment", "level": 2},
                        {"societal_goal": "Safe & healthy diets for all", "name": "Diet-related health conditions", "type": "Segment", "level": 2},
                        {"societal_goal": "Safe & healthy diets for all", "name": "Food safety and risk mitigation", "type": "Segment", "level": 2},
                        {"societal_goal": "Safe & healthy diets for all", "name": "Food integrity", "type": "Segment", "level": 2}
                    ]
                },
                {
                    "societal_goal": "Economically thriving, robust food value chains",
                    "name": "Economically thriving, robust food value chains",
                    "type": "Quadrant", 
                    "level": 1,
                    "children": [
                        {"societal_goal": "Economically thriving, robust food value chains", "name": "Innovative and transformative businesses", "type": "Segment", "level": 2},
                        {"societal_goal": "Economically thriving, robust food value chains", "name": "Dignified work environments", "type": "Segment", "level": 2},
                        {"societal_goal": "Economically thriving, robust food value chains", "name": "Value chain configuration", "type": "Segment", "level": 2},
                        {"societal_goal": "Economically thriving, robust food value chains", "name": "Competitive food systems", "type": "Segment", "level": 2}
                    ]
                },
                {
                    "societal_goal": "Clean & healthy planet",
                    "name": "Clean & healthy planet",
                    "type": "Quadrant",
                    "level": 1,
                    "children": [
                        {"societal_goal": "Clean & healthy planet", "name": "Biodiversity and ecosystem function", "type": "Segment", "level": 2},
                        {"societal_goal": "Clean & healthy planet", "name": "Primary production", "type": "Segment", "level": 2},
                        {"societal_goal": "Clean & healthy planet", "name": "Circularity", "type": "Segment", "level": 2},
                        {"societal_goal": "Clean & healthy planet", "name": "Climate change", "type": "Segment", "level": 2}
                    ]
                },
                {
                    "societal_goal": "Just, ethical, fair & culturally meaningful food system",
                    "name": "Just, ethical, fair & culturally meaningful food system",
                    "type": "Quadrant",
                    "level": 1,
                    "children": [
                        {"societal_goal": "Just, ethical, fair & culturally meaningful food system", "name": "Food access", "type": "Segment", "level": 2},
                        {"societal_goal": "Just, ethical, fair & culturally meaningful food system", "name": "Food environment", "type": "Segment", "level": 2},
                        {"societal_goal": "Just, ethical, fair & culturally meaningful food system", "name": "Environmental and social justice", "type": "Segment", "level": 2},
                        {"societal_goal": "Just, ethical, fair & culturally meaningful food system", "name": "Animal welfare", "type": "Segment", "level": 2}
                    ]
                }
            ]
            
            # Flatten the structure for Excel display
            excel_data = []
            for quadrant in food_compass_data:
                # Add quadrant row
                excel_data.append([quadrant["societal_goal"], quadrant["type"], quadrant["level"], ""])
                # Add segment rows
                for segment in quadrant["children"]:
                    excel_data.append([segment["societal_goal"], segment["type"], segment["level"], ""])
            
            headers = ["Societal goal (Quadrant)", "Type", "Level", "Description"]
            return jsonify({
                'success': True, 
                'headers': headers, 
                'data': excel_data,
                'total_rows': len(excel_data)
            })
        
        # Get all unique keys from all documents
        all_keys = set()
        for doc in documents:
            all_keys.update(doc.keys())
        
        # Convert to list and sort, but prioritize key fields first
        headers = sorted(list(all_keys))
        
        # Priority order for columns
        priority_fields = [
            'Quadrant',
            'Segment', 
            'Factor',
            'Location',
            'Indicator',
            'Title',
            'Url',
            'Datatype'
        ]
        
        # Fields to move to the end (usually MongoDB internal fields)
        end_fields = ['_id', '_created', '_updated', 'id']
        
        # Reorder headers according to priority
        ordered_headers = []
        for field in priority_fields:
            if field in headers:
                ordered_headers.append(field)
                headers.remove(field)
        
        # Add remaining headers in alphabetical order
        ordered_headers.extend(sorted(headers))
        
        # Move end fields to the back
        for field in end_fields:
            if field in ordered_headers:
                ordered_headers.remove(field)
                ordered_headers.append(field)
        
        headers = ordered_headers
        
        # Prepare data for Excel format
        excel_data = []
        for doc in documents:
            row = []
            for header in headers:
                value = doc.get(header, '')
                # Convert ObjectId to string
                if hasattr(value, '__str__'):
                    value = str(value)
                row.append(value)
            excel_data.append(row)
        
        return jsonify({
            'success': True, 
            'headers': headers, 
            'data': excel_data,
            'total_rows': len(excel_data)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/dataset/submit', methods=['POST'])
def submit_record():
    """Submit a new record to pending collection for review"""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})
        
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'})
        
        # Add submission timestamp
        data['submitted_at'] = str(datetime.now())
        
        # Insert the new document to pending collection
        result = pending_collection.insert_one(data)
        
        if result.inserted_id:
            return jsonify({
                'success': True, 
                'message': 'Record submitted for review',
                'inserted_id': str(result.inserted_id)
            })
        else:
            return jsonify({'success': False, 'error': 'Failed to submit record'})
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/dataset/pending')
def get_pending_items():
    """Get all pending items awaiting approval"""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})
        
        # Get all pending items
        items = list(pending_collection.find().sort('submitted_at', -1))
        
        # Convert ObjectId to string
        for item in items:
            if '_id' in item:
                item['_id'] = str(item['_id'])
        
        return jsonify({
            'success': True, 
            'items': items,
            'count': len(items)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/dataset/approve', methods=['POST'])
def approve_item():
    """Approve a pending item and move it to main collection"""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})
        
        data = request.get_json()
        if not data or 'item_id' not in data:
            return jsonify({'success': False, 'error': 'Item ID required'})
        
        item_id = data['item_id']
        
        # Get the pending item
        from bson import ObjectId
        pending_item = pending_collection.find_one({'_id': ObjectId(item_id)})
        
        if not pending_item:
            return jsonify({'success': False, 'error': 'Pending item not found'})
        
        # Remove submission timestamp and _id before adding to main collection
        item_data = {k: v for k, v in pending_item.items() if k not in ['_id', 'submitted_at']}
        
        # Insert into main collection
        result = collection.insert_one(item_data)
        
        if result.inserted_id:
            # Remove from pending collection
            pending_collection.delete_one({'_id': ObjectId(item_id)})
            
            return jsonify({
                'success': True, 
                'message': 'Item approved and added to database',
                'inserted_id': str(result.inserted_id)
            })
        else:
            return jsonify({'success': False, 'error': 'Failed to add item to database'})
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/dataset/reject', methods=['POST'])
def reject_item():
    """Reject a pending item and remove it from pending collection"""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})
        
        data = request.get_json()
        if not data or 'item_id' not in data:
            return jsonify({'success': False, 'error': 'Item ID required'})
        
        item_id = data['item_id']
        
        # Remove from pending collection
        from bson import ObjectId
        result = pending_collection.delete_one({'_id': ObjectId(item_id)})
        
        if result.deleted_count > 0:
            return jsonify({
                'success': True, 
                'message': 'Item rejected and removed'
            })
        else:
            return jsonify({'success': False, 'error': 'Item not found or already processed'})
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/dataset/add', methods=['POST'])
def add_record():
    """Add a new record to the collection (legacy endpoint)"""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})
        
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'})
        
        # Insert the new document
        result = collection.insert_one(data)
        
        if result.inserted_id:
            return jsonify({
                'success': True, 
                'message': 'Record added successfully',
                'inserted_id': str(result.inserted_id)
            })
        else:
            return jsonify({'success': False, 'error': 'Failed to insert record'})
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/auth/login', methods=['POST'])
def auth_login():
    """Authenticate user login"""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})
        
        data = request.get_json()
        if not data or 'username' not in data or 'password' not in data:
            return jsonify({'success': False, 'error': 'Username and password required'})
        
        username = data['username']
        password = data['password']
        
        # Check if user exists in Manager collection
        user = manage_collection.find_one({'Username': username})
        
        if not user:
            return jsonify({'success': False, 'error': 'Invalid username or password'})
        
        # Simple password check (in production, use proper password hashing)
        if str(user.get('Password')) != str(password):
            return jsonify({'success': False, 'error': 'Invalid username or password'})
        
        return jsonify({
            'success': True, 
            'message': 'Login successful',
            'username': username
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/auth/higher-manager-login', methods=['POST'])
def higher_manager_auth_login():
    """Authenticate higher manager login"""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})
        
        data = request.get_json()
        if not data or 'username' not in data or 'password' not in data:
            return jsonify({'success': False, 'error': 'Username and password required'})
        
        username = data['username']
        password = data['password']
        
        # Check if user exists in HigherManager collection
        user = higher_manager_collection.find_one({'Username': username})
        
        if not user:
            return jsonify({'success': False, 'error': 'Invalid higher manager credentials'})
        
        # Simple password check (in production, use proper password hashing)
        if str(user.get('Password')) != str(password):
            return jsonify({'success': False, 'error': 'Invalid higher manager credentials'})
        
        return jsonify({
            'success': True, 
            'message': 'Higher manager login successful',
            'username': username
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

def _build_global_or_query(goal, available_fields):
    """Build {$or: [...]} matching goal as substring (regex) across all non-internal fields."""
    or_conditions = []
    goal_lower = goal.lower().strip()
    if 'Location' in available_fields:
        if goal_lower == 'ireland':
            or_conditions.append({
                'Location': {"$regex": r'^Ireland$', "$options": "i"}
            })
        elif 'northern ireland' in goal_lower or goal_lower == 'n. ireland' or goal_lower == 'n ireland':
            or_conditions.append({
                'Location': {"$regex": r'^Northern Ireland$|^N\. Ireland$|^N Ireland$', "$options": "i"}
            })
        else:
            or_conditions.append({'Location': {"$regex": goal, "$options": "i"}})
    for fname in available_fields:
        if fname != 'Location':
            or_conditions.append({fname: {"$regex": goal, "$options": "i"}})
    if not or_conditions:
        return None
    return {"$or": or_conditions}

def _build_global_or_query_multi(keywords, available_fields):
    """Build {$or: [...]} matching ANY of the keywords across all fields (case-insensitive).

    This is used for multi-keyword search (we later score matches per-document in Python).
    """
    if not keywords:
        return None
    if isinstance(keywords, str):
        keywords = [keywords]
    kws = [str(k).strip() for k in keywords if str(k).strip()]
    if not kws:
        return None
    # Build a single safe regex alternation to avoid field_count * keyword_count explosion.
    if len(kws) == 1:
        pattern = re.escape(kws[0])
    else:
        pattern = r'(?:' + '|'.join(re.escape(k) for k in kws) + r')'
    or_conditions = []
    for fname in available_fields:
        or_conditions.append({fname: {"$regex": pattern, "$options": "i"}})
    return {"$or": or_conditions} if or_conditions else None


@app.route('/api/dataset/filter')
def filter_data():
    """Filter data: goal-only (global or scoped), or keyword-only (global), or keyword scoped to Quadrant/Segment/Factor."""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})

        goal = request.args.get('goal', '').strip()
        keyword = request.args.get('keyword', '').strip()
        field = request.args.get('field', '').strip()
        allowed_fields = {'Quadrant', 'Segment', 'Factor', 'Location'}
        quadrant = request.args.get('quadrant', '').strip()
        segment = request.args.get('segment', '').strip()
        factor = request.args.get('factor', '').strip()

        # Allow "scope-only" filtering (Quadrant/Segment/Factor) without requiring goal/keyword.
        # This is used by the table page when viewing "All Records".
        has_scope_only = bool(quadrant or segment or factor)
        if not goal and not keyword and not has_scope_only:
            return jsonify({'success': False, 'error': 'No goal, keyword, or scope filters provided'})

        def exact_value_regex(value: str):
            v = (value or '').strip()
            if not v:
                return None
            return {"$regex": rf'^\s*{re.escape(v)}\s*$', "$options": "i"}

        sample_docs = list(collection.find().limit(5))
        available_fields = set()
        for doc in sample_docs:
            available_fields.update(doc.keys())
        available_fields.discard('_id')
        available_fields.discard('_created')
        available_fields.discard('_updated')
        available_fields.discard('id')

        # --- Scope-only filtering (no goal and no keyword) ---
        if not goal and not keyword and has_scope_only:
            scope_parts = []
            if quadrant and 'Quadrant' in available_fields:
                rx = exact_value_regex(quadrant)
                if rx:
                    scope_parts.append({'Quadrant': rx})
            if segment and 'Segment' in available_fields:
                rx = exact_value_regex(segment)
                if rx:
                    scope_parts.append({'Segment': rx})
            if factor and 'Factor' in available_fields:
                rx = exact_value_regex(factor)
                if rx:
                    scope_parts.append({'Factor': rx})

            query = {'$and': scope_parts} if len(scope_parts) > 1 else (scope_parts[0] if scope_parts else {})
            documents = list(collection.find(query))
            for doc in documents:
                if '_id' in doc:
                    doc['_id'] = str(doc['_id'])
            return jsonify({
                'success': True,
                'records': documents,
                'count': len(documents),
                'query_used': query,
                'available_fields': list(available_fields)
            })

        # --- Keyword search: global OR scoped to current Factor/Segment/Quadrant ---
        if keyword:
            def split_keywords(raw: str):
                """Split raw keyword string into multiple keywords.

                - If the user uses comma/Chinese comma/semicolon/newline separators, split on those
                  (so phrases like "food safety" can remain intact).
                - Otherwise, split on whitespace.
                """
                s = (raw or '').strip()
                if not s:
                    return []
                # Normalize separators to commas
                for sep in ['，', ';', '；', '\n', '\t', '|']:
                    s = s.replace(sep, ',')
                if ',' in s:
                    parts = [p.strip() for p in s.split(',')]
                else:
                    parts = re.split(r'\s+', s)
                out = []
                seen = set()
                for p in parts:
                    if not p:
                        continue
                    key = p.strip().lower()
                    if not key or key in seen:
                        continue
                    seen.add(key)
                    out.append(p.strip())
                return out

            keywords = split_keywords(keyword)
            # Backward-compatible: if parsing yields nothing, treat as a single keyword.
            if not keywords:
                keywords = [keyword]

            keyword_query = (
                _build_global_or_query(keywords[0], available_fields)
                if len(keywords) == 1
                else _build_global_or_query_multi(keywords, available_fields)
            )
            if not keyword_query:
                return jsonify({
                    'success': False,
                    'error': f'No matching fields found. Available fields: {list(available_fields)}',
                    'available_fields': list(available_fields),
                    'sample_doc': sample_docs[0] if sample_docs else None
                })

            has_scope = (
                (field and field in allowed_fields and field in available_fields and bool(goal))
                or bool(quadrant or segment or factor)
            )

            if not has_scope:
                query = keyword_query
            else:
                scope_parts = []
                if field and field in allowed_fields and field in available_fields and goal:
                    goal_lower = goal.lower().strip()
                    if field == 'Location':
                        if goal_lower == 'ireland':
                            scope_parts.append({'Location': {"$regex": r'^Ireland$', "$options": "i"}})
                        elif 'northern ireland' in goal_lower or goal_lower == 'n. ireland' or goal_lower == 'n ireland':
                            scope_parts.append({'Location': {"$regex": r'^Northern Ireland$|^N\. Ireland$|^N Ireland$', "$options": "i"}})
                        elif goal_lower in ['united kingdom', 'uk']:
                            scope_parts.append({'Location': {"$regex": r'^(United Kingdom|UK)$', "$options": "i"}})
                        else:
                            scope_parts.append({'Location': {"$regex": f'^{re.escape(goal.strip())}$', "$options": "i"}})
                    else:
                        scope_parts.append({field: {"$regex": f'^{re.escape(goal.strip())}$', "$options": "i"}})
                if quadrant:
                    rx = exact_value_regex(quadrant)
                    scope_parts.append({'Quadrant': rx} if rx else {'Quadrant': quadrant})
                if segment:
                    rx = exact_value_regex(segment)
                    scope_parts.append({'Segment': rx} if rx else {'Segment': segment})
                if factor:
                    rx = exact_value_regex(factor)
                    scope_parts.append({'Factor': rx} if rx else {'Factor': factor})

                if not scope_parts:
                    query = keyword_query
                elif len(scope_parts) == 1:
                    query = {"$and": [scope_parts[0], keyword_query]}
                else:
                    query = {"$and": [{'$and': scope_parts}, keyword_query]}

            documents = list(collection.find(query))

            # Rank results by how many keywords they match (best match first).
            # We do this in Python to support "best match" behaviour without complex aggregation.
            def doc_to_search_text(doc):
                chunks = []
                for fname in available_fields:
                    if fname not in doc:
                        continue
                    v = doc.get(fname)
                    if v is None:
                        continue
                    if isinstance(v, (dict, list)):
                        try:
                            chunks.append(json.dumps(v, ensure_ascii=False))
                        except Exception:
                            chunks.append(str(v))
                    else:
                        chunks.append(str(v))
                return '\n'.join(chunks).lower()

            scored = []
            for doc in documents:
                text = doc_to_search_text(doc)
                match_count = 0
                matched = []
                for kw in keywords:
                    if kw.lower() in text:
                        match_count += 1
                        matched.append(kw)
                scored.append((match_count, doc))

            scored.sort(key=lambda t: t[0], reverse=True)
            documents = [d for (score, d) in scored if score > 0]
            best_match_count = scored[0][0] if scored else 0
            full_match_count = sum(1 for (score, _d) in scored if score == len(keywords)) if keywords else 0
            match_mode = (
                'none' if not documents
                else ('all_keywords' if full_match_count > 0 else 'partial_best_match')
            )

            # Provide per-record match info without adding extra columns to records.
            scores_by_id = {}
            for (score, doc) in scored:
                if score <= 0:
                    continue
                try:
                    doc_id = str(doc.get('_id'))
                except Exception:
                    doc_id = None
                if not doc_id:
                    continue
                text = doc_to_search_text(doc)
                matched = []
                for kw in keywords:
                    if kw.lower() in text:
                        matched.append(kw)
                scores_by_id[doc_id] = {
                    'match_count': int(score),
                    'matched_keywords': matched
                }

            for doc in documents:
                if '_id' in doc:
                    doc['_id'] = str(doc['_id'])
            return jsonify({
                'success': True,
                'records': documents,
                'count': len(documents),
                'query_used': query,
                'available_fields': list(available_fields),
                'scoped_keyword_search': bool(keyword and has_scope),
                'keyword_search': {
                    'keywords': keywords,
                    'total_keywords': len(keywords),
                    'best_match_count': int(best_match_count),
                    'full_match_count': int(full_match_count),
                    'mode': match_mode,
                    'scores_by_id': scores_by_id
                }
            })

        # --- No keyword: original goal-based behaviour ---
        or_conditions = []
        goal_lower = goal.lower().strip()
        is_location_search = (
            goal_lower in ['ireland', 'northern ireland', 'united kingdom', 'uk']
            or goal_lower == 'n. ireland'
            or goal_lower == 'n ireland'
            or 'northern ireland' in goal_lower
        )

        if field and field in allowed_fields and field in available_fields:
            if field == 'Location':
                if goal_lower == 'ireland':
                    query = {'Location': {"$regex": r'^Ireland$', "$options": "i"}}
                elif 'northern ireland' in goal_lower or goal_lower == 'n. ireland' or goal_lower == 'n ireland':
                    query = {'Location': {"$regex": r'^Northern Ireland$|^N\. Ireland$|^N Ireland$', "$options": "i"}}
                elif goal_lower in ['united kingdom', 'uk']:
                    query = {'Location': {"$regex": r'^(United Kingdom|UK)$', "$options": "i"}}
                else:
                    query = {'Location': {"$regex": f'^{re.escape(goal.strip())}$', "$options": "i"}}
            else:
                query = {field: {"$regex": f'^{re.escape(goal.strip())}$', "$options": "i"}}

        elif is_location_search and 'Location' in available_fields:
            if goal_lower == 'ireland':
                query = {'Location': {"$regex": r'^Ireland$', "$options": "i"}}
            elif 'northern ireland' in goal_lower or goal_lower == 'n. ireland' or goal_lower == 'n ireland':
                query = {'Location': {"$regex": r'^Northern Ireland$|^N\. Ireland$|^N Ireland$', "$options": "i"}}
            elif goal_lower in ['united kingdom', 'uk']:
                query = {'Location': {"$regex": r'^(United Kingdom|UK)$', "$options": "i"}}
            else:
                query = {'Location': {"$regex": f'^{re.escape(goal.strip())}$', "$options": "i"}}
        else:
            if 'Location' in available_fields:
                if goal_lower == 'ireland':
                    or_conditions.append({'Location': {"$regex": r'^Ireland$', "$options": "i"}})
                elif 'northern ireland' in goal_lower or goal_lower == 'n. ireland' or goal_lower == 'n ireland':
                    or_conditions.append({'Location': {"$regex": r'^Northern Ireland$|^N\. Ireland$|^N Ireland$', "$options": "i"}})
                else:
                    or_conditions.append({'Location': {"$regex": goal, "$options": "i"}})
            for fname in available_fields:
                if fname != 'Location':
                    or_conditions.append({fname: {"$regex": goal, "$options": "i"}})
            if not or_conditions:
                return jsonify({
                    'success': False,
                    'error': f'No matching fields found. Available fields: {list(available_fields)}',
                    'available_fields': list(available_fields),
                    'sample_doc': sample_docs[0] if sample_docs else None
                })
            query = {"$or": or_conditions}

        additional_filters = {}
        if quadrant:
            additional_filters['Quadrant'] = quadrant
        if segment:
            additional_filters['Segment'] = segment
        if factor:
            additional_filters['Factor'] = factor

        if additional_filters:
            if isinstance(query, dict) and '$or' in query:
                query = {"$and": [query, additional_filters]}
            else:
                query = {**query, **additional_filters}

        documents = list(collection.find(query))
        for doc in documents:
            if '_id' in doc:
                doc['_id'] = str(doc['_id'])
        return jsonify({
            'success': True,
            'records': documents,
            'count': len(documents),
            'query_used': query,
            'available_fields': list(available_fields)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/descriptions')
def get_descriptions():
    """Get module descriptions from database for tooltips"""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})
        
        # Fetch all documents from description collection
        # Structure: { keywords: [{ name: string, short_description: string, definition: string }] }
        documents = list(description_collection.find({}, {'_id': 0}))
        
        # Convert to dictionary for easy lookup by name
        # Extract all keywords from all documents and flatten into a single dictionary
        descriptions_dict = {}
        for doc in documents:
            if 'keywords' in doc and isinstance(doc['keywords'], list):
                for keyword in doc['keywords']:
                    if 'name' in keyword:
                        descriptions_dict[keyword['name']] = {
                            'short_description': keyword.get('short_description', ''),
                            'definition': keyword.get('definition', '')
                        }
        
        return jsonify({'success': True, 'descriptions': descriptions_dict})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/dataset/export-filtered')
def export_filtered_data():
    """Export filtered data as Excel file"""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})
        
        # Get the goal parameter from query string
        goal = request.args.get('goal', '')
        keyword = request.args.get('keyword', '').strip()
        goal_label = goal.strip() or (keyword.strip()[:60] if keyword else 'ALL')
        # Optional: restrict matching to a specific field (Quadrant/Segment/Factor/Location)
        field = request.args.get('field', '').strip()
        allowed_fields = {'Quadrant', 'Segment', 'Factor', 'Location'}

        # Optional additional filters
        quadrant = request.args.get('quadrant', '').strip()
        segment = request.args.get('segment', '').strip()
        factor = request.args.get('factor', '').strip()

        has_scope_only = bool(quadrant or segment or factor)
        if not goal and not keyword and not has_scope_only:
            return jsonify({'success': False, 'error': 'No goal, keyword, or scope filters provided'})
        
        # Get all available fields from sample documents
        sample_docs = list(collection.find().limit(5))
        available_fields = set()
        for doc in sample_docs:
            available_fields.update(doc.keys())
        
        # Remove MongoDB internal fields
        available_fields.discard('_id')
        available_fields.discard('_created')
        available_fields.discard('_updated')
        available_fields.discard('id')

        def exact_value_regex(value: str):
            v = (value or '').strip()
            if not v:
                return None
            return {"$regex": rf'^\s*{re.escape(v)}\s*$', "$options": "i"}

        # Scope-only export (no goal/keyword): just export rows matching Quadrant/Segment/Factor filters
        if not goal and not keyword and has_scope_only:
            parts = []
            if quadrant and 'Quadrant' in available_fields:
                rx = exact_value_regex(quadrant)
                if rx:
                    parts.append({'Quadrant': rx})
            if segment and 'Segment' in available_fields:
                rx = exact_value_regex(segment)
                if rx:
                    parts.append({'Segment': rx})
            if factor and 'Factor' in available_fields:
                rx = exact_value_regex(factor)
                if rx:
                    parts.append({'Factor': rx})

            query = {'$and': parts} if len(parts) > 1 else (parts[0] if parts else {})
            documents = list(collection.find(query))
            # Continue to the common Excel generation code below using `documents`
        else:
            documents = None  # will be computed by existing logic below
        
        if documents is None:
            # --- Keyword export: same behaviour as keyword filtering on the table page ---
            if keyword:
                def split_keywords(raw: str):
                    s = (raw or '').strip()
                    if not s:
                        return []
                    for sep in ['，', ';', '；', '\n', '\t', '|']:
                        s = s.replace(sep, ',')
                    if ',' in s:
                        parts = [p.strip() for p in s.split(',')]
                    else:
                        parts = re.split(r'\s+', s)
                    out = []
                    seen = set()
                    for p in parts:
                        if not p:
                            continue
                        key = p.strip().lower()
                        if not key or key in seen:
                            continue
                        seen.add(key)
                        out.append(p.strip())
                    return out

                keywords = split_keywords(keyword) or [keyword]

                # Build a query that matches ANY keyword across ANY field, then rank in Python.
                if len(keywords) == 1:
                    keyword_query = _build_global_or_query(keywords[0], available_fields)
                else:
                    keyword_query = _build_global_or_query_multi(keywords, available_fields)
                if not keyword_query:
                    return jsonify({'success': False, 'error': 'No searchable fields found for keyword export'})

                has_scope = (
                    (field and field in allowed_fields and field in available_fields and bool(goal))
                    or bool(quadrant or segment or factor)
                )

                if not has_scope:
                    query = keyword_query
                else:
                    scope_parts = []
                    # Optional goal+field scoping (module lock)
                    if field and field in allowed_fields and field in available_fields and goal:
                        goal_lower = goal.lower().strip()
                        if field == 'Location':
                            if goal_lower == 'ireland':
                                scope_parts.append({'Location': {"$regex": r'^Ireland$', "$options": "i"}})
                            elif 'northern ireland' in goal_lower or goal_lower == 'n. ireland' or goal_lower == 'n ireland':
                                scope_parts.append({'Location': {"$regex": r'^Northern Ireland$|^N\. Ireland$|^N Ireland$', "$options": "i"}})
                            elif goal_lower in ['united kingdom', 'uk']:
                                scope_parts.append({'Location': {"$regex": r'^(United Kingdom|UK)$', "$options": "i"}})
                            else:
                                scope_parts.append({'Location': {"$regex": f'^{re.escape(goal.strip())}$', "$options": "i"}})
                        else:
                            scope_parts.append({field: {"$regex": f'^{re.escape(goal.strip())}$', "$options": "i"}})

                    # Optional Q/S/F filters
                    if quadrant:
                        rx = exact_value_regex(quadrant)
                        scope_parts.append({'Quadrant': rx} if rx else {'Quadrant': quadrant})
                    if segment:
                        rx = exact_value_regex(segment)
                        scope_parts.append({'Segment': rx} if rx else {'Segment': segment})
                    if factor:
                        rx = exact_value_regex(factor)
                        scope_parts.append({'Factor': rx} if rx else {'Factor': factor})

                    if not scope_parts:
                        query = keyword_query
                    elif len(scope_parts) == 1:
                        query = {"$and": [scope_parts[0], keyword_query]}
                    else:
                        query = {"$and": [{'$and': scope_parts}, keyword_query]}

                raw_docs = list(collection.find(query))

                def doc_to_search_text(doc):
                    chunks = []
                    for fname in available_fields:
                        if fname not in doc:
                            continue
                        v = doc.get(fname)
                        if v is None:
                            continue
                        if isinstance(v, (dict, list)):
                            try:
                                chunks.append(json.dumps(v, ensure_ascii=False))
                            except Exception:
                                chunks.append(str(v))
                        else:
                            chunks.append(str(v))
                    return '\n'.join(chunks).lower()

                scored = []
                for doc in raw_docs:
                    text = doc_to_search_text(doc)
                    match_count = 0
                    for kw in keywords:
                        if kw.lower() in text:
                            match_count += 1
                    if match_count > 0:
                        scored.append((match_count, doc))

                scored.sort(key=lambda t: t[0], reverse=True)
                documents = [d for (_s, d) in scored]

            # --- No keyword: original goal-based behaviour ---
            # Create search query for all fields
            search_conditions = []
            goal_lower = goal.lower().strip()
            
            # Known location values in the database
            known_locations = ['ireland', 'northern ireland', 'united kingdom', 'uk']
            
            # Check if the goal is a known location - if so, only search Location field
            is_location_search = goal_lower in known_locations or \
                               goal_lower == 'n. ireland' or \
                               goal_lower == 'n ireland' or \
                               'northern ireland' in goal_lower
            
            # If a specific field is requested, do a field-specific exact match (case-insensitive)
            if field and field in allowed_fields and field in available_fields:
                if field == 'Location':
                    if goal_lower == 'ireland':
                        query = {'Location': {"$regex": r'^Ireland$', "$options": "i"}}
                    elif 'northern ireland' in goal_lower or goal_lower == 'n. ireland' or goal_lower == 'n ireland':
                        query = {'Location': {"$regex": r'^Northern Ireland$|^N\. Ireland$|^N Ireland$', "$options": "i"}}
                    elif goal_lower in ['united kingdom', 'uk']:
                        query = {'Location': {"$regex": r'^(United Kingdom|UK)$', "$options": "i"}}
                    else:
                        query = {'Location': {"$regex": f'^{re.escape(goal.strip())}$', "$options": "i"}}
                else:
                    query = {field: {"$regex": f'^{re.escape(goal.strip())}$', "$options": "i"}}

            elif is_location_search and 'Location' in available_fields:
                # For location searches, ONLY search in Location field with exact match
                if goal_lower == 'ireland':
                    # For "Ireland", match exactly "Ireland" but exclude "Northern Ireland"
                    query = {
                        'Location': {
                            "$regex": r'^Ireland$',
                            "$options": "i"
                        }
                    }
                elif 'northern ireland' in goal_lower or goal_lower == 'n. ireland' or goal_lower == 'n ireland':
                    # For "Northern Ireland", match only "Northern Ireland" variations
                    query = {
                        'Location': {
                            "$regex": r'^Northern Ireland$|^N\. Ireland$|^N Ireland$',
                            "$options": "i"
                        }
                    }
                elif goal_lower in ['united kingdom', 'uk']:
                    # For "United Kingdom" or "UK", match exactly
                    query = {
                        'Location': {
                            "$regex": r'^(United Kingdom|UK)$',
                            "$options": "i"
                        }
                    }
                else:
                    # Fallback: exact match for location
                    query = {
                        'Location': {
                            "$regex": f'^{re.escape(goal.strip())}$',
                            "$options": "i"
                        }
                    }
            else:
                # For non-location searches, search ALL fields
                # Special handling for Location field to distinguish Ireland from Northern Ireland
                if 'Location' in available_fields:
                    if goal_lower == 'ireland':
                        # For "Ireland", match exactly "Ireland" but exclude "Northern Ireland"
                        search_conditions.append({
                            'Location': {
                                "$regex": r'^Ireland$',
                                "$options": "i"
                            }
                        })
                    elif 'northern ireland' in goal_lower or goal_lower == 'n. ireland' or goal_lower == 'n ireland':
                        # For "Northern Ireland", match only "Northern Ireland" variations
                        search_conditions.append({
                            'Location': {
                                "$regex": r'^Northern Ireland$|^N\. Ireland$|^N Ireland$',
                                "$options": "i"
                            }
                        })
                    else:
                        # For other locations, use normal regex matching
                        search_conditions.append({'Location': {"$regex": goal, "$options": "i"}})
                
                # For all other fields, use normal regex matching
                for field in available_fields:
                    if field != 'Location':  # Location already handled above
                        search_conditions.append({field: {"$regex": goal, "$options": "i"}})
                
                query = {"$or": search_conditions} if search_conditions else {}

            # Add additional filters (quadrant, segment, factor) if provided
            additional_filters = {}
            if quadrant:
                additional_filters['Quadrant'] = quadrant
            if segment:
                additional_filters['Segment'] = segment
            if factor:
                additional_filters['Factor'] = factor

            if additional_filters:
                if isinstance(query, dict) and '$or' in query:
                    query = {"$and": [query, additional_filters]}
                else:
                    query.update(additional_filters)

            if documents is None:
                documents = list(collection.find(query))
        
        if not documents:
            return jsonify({'success': False, 'error': 'No data found for the specified filter'})
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Filtered Data - {goal_label}"
        
        # Get all unique keys from all documents
        all_keys = set()
        for doc in documents:
            all_keys.update(doc.keys())
        
        # Priority order for columns
        priority_fields = [
            'Quadrant',
            'Segment', 
            'Factor',
            'Location',
            'Indicator',
            'Title',
            'Url',
            'Datatype'
        ]
        
        # Fields to move to the end
        end_fields = ['_id', '_created', '_updated', 'id']
        
        # Reorder headers according to priority
        ordered_headers = []
        remaining_keys = list(all_keys)
        
        for field in priority_fields:
            if field in remaining_keys:
                ordered_headers.append(field)
                remaining_keys.remove(field)
        
        # Add remaining headers in alphabetical order
        ordered_headers.extend(sorted(remaining_keys))
        
        # Move end fields to the back
        for field in end_fields:
            if field in ordered_headers:
                ordered_headers.remove(field)
                ordered_headers.append(field)
        
        headers = ordered_headers
        
        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Write data
        for row, doc in enumerate(documents, 2):
            for col, header in enumerate(headers, 1):
                value = doc.get(header, '')
                # Convert ObjectId to string
                if hasattr(value, '__str__'):
                    value = str(value)
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, len(documents) + 2):
                cell_value = ws[f"{column_letter}{row}"].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=filtered_data_{goal_label.replace(" ", "_")}.xlsx'}
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/dataset/export')
def export_data():
    """Export data as Excel file"""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})
        
        documents = list(collection.find())
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Dataset"
        
        if documents:
            # Get all unique keys from all documents
            all_keys = set()
            for doc in documents:
                all_keys.update(doc.keys())
            
            # Convert to list and sort, but prioritize key fields first
            headers = sorted(list(all_keys))
            
        # Priority order for columns
        priority_fields = [
            'Quadrant',
            'Segment', 
            'Factor',
            'Location',
            'Indicator',
            'Title',
            'Url',
            'Datatype'
        ]
        
        # Fields to move to the end (usually MongoDB internal fields)
        end_fields = ['_id', '_created', '_updated', 'id']
        
        # Reorder headers according to priority
        ordered_headers = []
        for field in priority_fields:
            if field in headers:
                ordered_headers.append(field)
                headers.remove(field)
        
        # Add remaining headers in alphabetical order
        ordered_headers.extend(sorted(headers))
        
        # Move end fields to the back
        for field in end_fields:
            if field in ordered_headers:
                ordered_headers.remove(field)
                ordered_headers.append(field)
        
        headers = ordered_headers
        
        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Write data
        for row, doc in enumerate(documents, 2):
            for col, header in enumerate(headers, 1):
                value = doc.get(header, '')
                # Convert ObjectId to string
                if hasattr(value, '__str__'):
                    value = str(value)
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, len(documents) + 2):
                cell_value = ws[f"{column_letter}{row}"].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=dataset_export.xlsx'}
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/higher-manager-dashboard')
def higher_manager_dashboard():
    """Higher Manager Dashboard page"""
    return render_template('higher_manager_dashboard.html')

@app.route('/manager')
def manager():
    """Serve the database management tool"""
    return render_template('manager.html')

@app.route('/api/dataset/all')
def get_all_data():
    """Get all data from the main collection"""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})
        
        documents = list(collection.find())
        
        if not documents:
            return jsonify({'success': False, 'error': 'No data found in database'})
        
        # Convert ObjectId to string for JSON serialization
        for doc in documents:
            if '_id' in doc:
                doc['_id'] = str(doc['_id'])
        
        return jsonify({
            'success': True,
            'records': documents,
            'count': len(documents)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/dataset/add-direct', methods=['POST'])
def add_direct_record():
    """Add a record directly to the main collection (Higher Manager only)"""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})
        
        data = request.get_json()
        
        # Remove _id if present (MongoDB will generate new one)
        if '_id' in data:
            del data['_id']
        
        # Add timestamp
        data['_created'] = datetime.now()
        data['_updated'] = datetime.now()
        
        # Insert into main collection
        result = collection.insert_one(data)
        
        return jsonify({
            'success': True,
            'message': 'Record added successfully',
            'inserted_id': str(result.inserted_id)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/dataset/update', methods=['PUT'])
def update_record():
    """Update a record in the main collection (Higher Manager only)"""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})
        
        data = request.get_json()
        record_id = data.get('_id')
        
        if not record_id:
            return jsonify({'success': False, 'error': 'Record ID is required'})
        
        # Remove _id from update data
        del data['_id']
        
        # Add update timestamp
        data['_updated'] = datetime.now()
        
        # Update the record
        result = collection.update_one(
            {'_id': ObjectId(record_id)},
            {'$set': data}
        )
        
        if result.matched_count == 0:
            return jsonify({'success': False, 'error': 'Record not found'})
        
        return jsonify({
            'success': True,
            'message': 'Record updated successfully'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/dataset/delete', methods=['DELETE'])
def delete_record():
    """Delete a record from the main collection (Higher Manager only)"""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})
        
        data = request.get_json()
        record_id = data.get('_id')
        
        if not record_id:
            return jsonify({'success': False, 'error': 'Record ID is required'})
        
        # Delete the record
        result = collection.delete_one({'_id': ObjectId(record_id)})
        
        if result.deleted_count == 0:
            return jsonify({'success': False, 'error': 'Record not found'})
        
        return jsonify({
            'success': True,
            'message': 'Record deleted successfully'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/dataset/export-all')
def export_all_data():
    """Export all data to Excel file"""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})
        
        documents = list(collection.find())
        
        if not documents:
            return jsonify({'success': False, 'error': 'No data found to export'})
        
        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "All Database Records"
        
        # Get all unique headers
        all_headers = set()
        for doc in documents:
            all_headers.update(doc.keys())
        
        # Remove MongoDB internal fields
        all_headers.discard('_id')
        all_headers.discard('_created')
        all_headers.discard('_updated')
        all_headers.discard('id')
        
        # Priority order for columns
        priority_fields = [
            'Quadrant', 'Segment', 'Factor', 'Location', 
            'Indicator', 'Title', 'Url', 'Datatype'
        ]
        
        # Fields to move to the end
        end_fields = ['_id', '_created', '_updated', 'id']
        
        # Reorder headers according to priority
        ordered_headers = []
        headers = list(all_headers)
        
        for field in priority_fields:
            if field in headers:
                ordered_headers.append(field)
                headers.remove(field)
        
        # Add remaining headers in alphabetical order
        ordered_headers.extend(sorted(headers))
        
        # Move end fields to the back
        for field in end_fields:
            if field in ordered_headers:
                ordered_headers.remove(field)
                ordered_headers.append(field)
        
        headers = ordered_headers
        
        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Write data
        for row, doc in enumerate(documents, 2):
            for col, header in enumerate(headers, 1):
                value = doc.get(header, '')
                # Convert ObjectId to string
                if hasattr(value, '__str__'):
                    value = str(value)
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, len(documents) + 2):
                cell_value = ws[f"{column_letter}{row}"].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=all_database_records.xlsx'}
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/map')
def map():
    """Serve the location map visualization page"""
    return render_template('map.html')

@app.route('/api/dataset/location-counts')
def get_location_counts():
    """Get count of items per location, with optional filters for Quadrant, Segment, Factor"""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})
        
        # Get filter parameters from query string
        quadrant = request.args.get('quadrant', '').strip()
        segment = request.args.get('segment', '').strip()
        factor = request.args.get('factor', '').strip()
        
        # Build match conditions
        match_conditions = {
            'Location': {'$exists': True, '$ne': None, '$ne': ''}
        }
        
        # Add filter conditions if provided
        if quadrant:
            match_conditions['Quadrant'] = quadrant
        if segment:
            match_conditions['Segment'] = segment
        if factor:
            match_conditions['Factor'] = factor
        
        # Get all documents with Location field and apply filters
        pipeline = [
            {
                '$match': match_conditions
            },
            {
                '$group': {
                    '_id': '$Location',
                    'count': {'$sum': 1}
                }
            },
            {
                '$sort': {'count': -1}
            }
        ]
        
        location_counts = list(collection.aggregate(pipeline))
        
        # Format the results
        result = []
        for item in location_counts:
            result.append({
                'location': item['_id'],
                'count': item['count']
            })
        
        return jsonify({
            'success': True,
            'data': result
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/dataset/unique-values')
def get_unique_values():
    """Get unique values for Quadrant, Segment, Factor for cascading dropdowns"""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})
        
        # Get unique values for each field
        quadrants_raw = list(collection.distinct('Quadrant'))
        segments_raw = list(collection.distinct('Segment'))
        factors_raw = list(collection.distinct('Factor'))
        locations_raw = list(collection.distinct('Location'))
        
        # Normalize and deduplicate values
        # For Quadrant, Segment, Factor: strip whitespace, normalize case, remove duplicates
        def normalize_and_deduplicate(values):
            # First, normalize each value (strip whitespace, but keep original case for display)
            normalized_map = {}  # Maps normalized value to original value
            for v in values:
                if v:  # Skip None/empty values
                    # Strip leading/trailing whitespace
                    v_stripped = str(v).strip()
                    if v_stripped:
                        # Use lowercase for comparison, but keep original for display
                        v_normalized = v_stripped.lower()
                        # If we haven't seen this normalized value, or if the current one is shorter (prefer shorter)
                        if v_normalized not in normalized_map or len(v_stripped) < len(normalized_map[v_normalized]):
                            normalized_map[v_normalized] = v_stripped
            # Return sorted unique values (using original case)
            return sorted(set(normalized_map.values()))
        
        quadrants = normalize_and_deduplicate(quadrants_raw)
        segments = normalize_and_deduplicate(segments_raw)
        factors = normalize_and_deduplicate(factors_raw)
        locations = normalize_and_deduplicate(locations_raw)
        
        # Debug: Log the values to help identify duplicates
        print(f"DEBUG: Raw Quadrants ({len(quadrants_raw)}): {quadrants_raw}")
        print(f"DEBUG: Normalized Quadrants ({len(quadrants)}): {quadrants}")
        
        # Create mapping for reverse lookup
        quadrant_segment_map = {}
        segment_factor_map = {}
        factor_segment_map = {}
        factor_quadrant_map = {}
        segment_quadrant_map = {}  # Add segment to quadrant mapping
        
        # Build mappings
        for doc in collection.find({}, {'Quadrant': 1, 'Segment': 1, 'Factor': 1}):
            quadrant = doc.get('Quadrant')
            segment = doc.get('Segment')
            factor = doc.get('Factor')
            
            if quadrant and segment:
                if quadrant not in quadrant_segment_map:
                    quadrant_segment_map[quadrant] = set()
                quadrant_segment_map[quadrant].add(segment)
                
                # Add reverse mapping: segment to quadrant
                if segment not in segment_quadrant_map:
                    segment_quadrant_map[segment] = set()
                segment_quadrant_map[segment].add(quadrant)
            
            if segment and factor:
                if segment not in segment_factor_map:
                    segment_factor_map[segment] = set()
                segment_factor_map[segment].add(factor)
            
            if factor and segment:
                if factor not in factor_segment_map:
                    factor_segment_map[factor] = set()
                factor_segment_map[factor].add(segment)
            
            if factor and quadrant:
                if factor not in factor_quadrant_map:
                    factor_quadrant_map[factor] = set()
                factor_quadrant_map[factor].add(quadrant)
        
        # Convert sets to sorted lists
        for key in quadrant_segment_map:
            quadrant_segment_map[key] = sorted(list(quadrant_segment_map[key]))
        for key in segment_factor_map:
            segment_factor_map[key] = sorted(list(segment_factor_map[key]))
        for key in factor_segment_map:
            factor_segment_map[key] = sorted(list(factor_segment_map[key]))
        for key in factor_quadrant_map:
            factor_quadrant_map[key] = sorted(list(factor_quadrant_map[key]))
        for key in segment_quadrant_map:
            segment_quadrant_map[key] = sorted(list(segment_quadrant_map[key]))
        
        return jsonify({
            'success': True,
            'data': {
                'quadrants': quadrants,
                'segments': segments,
                'factors': factors,
                'locations': locations,
                'mappings': {
                    'quadrant_segment': quadrant_segment_map,
                    'segment_factor': segment_factor_map,
                    'factor_segment': factor_segment_map,
                    'factor_quadrant': factor_quadrant_map,
                    'segment_quadrant': segment_quadrant_map
                }
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/compass/structure')
def get_compass_structure():
    """Get Compass structure (Quadrant, Segment, Factor) from Compass collection"""
    try:
        if not mongo_connected:
            return jsonify({'success': False, 'error': 'MongoDB not connected'})
        
        # Get all documents from Compass collection
        compass_docs = list(compass_collection.find({}))
        
        if not compass_docs:
            # If collection is empty, try to load from compass.json file
            compass_file_path = os.path.join(os.path.dirname(__file__), 'compass.json')
            if os.path.exists(compass_file_path):
                with open(compass_file_path, 'r', encoding='utf-8') as f:
                    compass_data = json.load(f)
                    # Insert into collection if empty
                    if compass_data:
                        compass_collection.insert_many(compass_data)
                        compass_docs = list(compass_collection.find({}))
        
        # Build structure from Compass collection
        quadrants = []
        segments = []
        factors = []
        
        # Mappings for cascading filters
        quadrant_segment_map = {}
        segment_factor_map = {}
        # Reverse mappings for auto-fill
        segment_quadrant_map = {}
        factor_segment_map = {}
        factor_quadrant_map = {}
        
        for doc in compass_docs:
            quadrant = doc.get('Quadrant')
            if not quadrant:
                continue
            
            quadrants.append(quadrant)
            children = doc.get('children', [])
            
            if quadrant not in quadrant_segment_map:
                quadrant_segment_map[quadrant] = []
            
            for child in children:
                segment = child.get('Segment')
                if not segment:
                    continue
                
                segments.append(segment)
                quadrant_segment_map[quadrant].append(segment)
                
                # Reverse mapping: segment to quadrant
                if segment not in segment_quadrant_map:
                    segment_quadrant_map[segment] = []
                if quadrant not in segment_quadrant_map[segment]:
                    segment_quadrant_map[segment].append(quadrant)
                
                factor_list = child.get('Factor', [])
                if segment not in segment_factor_map:
                    segment_factor_map[segment] = []
                
                for factor in factor_list:
                    if factor:
                        factors.append(factor)
                        segment_factor_map[segment].append(factor)
                        
                        # Reverse mapping: factor to segment
                        if factor not in factor_segment_map:
                            factor_segment_map[factor] = []
                        if segment not in factor_segment_map[factor]:
                            factor_segment_map[factor].append(segment)
                        
                        # Reverse mapping: factor to quadrant
                        if factor not in factor_quadrant_map:
                            factor_quadrant_map[factor] = []
                        if quadrant not in factor_quadrant_map[factor]:
                            factor_quadrant_map[factor].append(quadrant)
        
        # Remove duplicates and sort
        quadrants = sorted(list(set(quadrants)))
        segments = sorted(list(set(segments)))
        factors = sorted(list(set(factors)))
        
        # Sort mappings
        for key in quadrant_segment_map:
            quadrant_segment_map[key] = sorted(list(set(quadrant_segment_map[key])))
        for key in segment_factor_map:
            segment_factor_map[key] = sorted(list(set(segment_factor_map[key])))
        for key in segment_quadrant_map:
            segment_quadrant_map[key] = sorted(list(set(segment_quadrant_map[key])))
        for key in factor_segment_map:
            factor_segment_map[key] = sorted(list(set(factor_segment_map[key])))
        for key in factor_quadrant_map:
            factor_quadrant_map[key] = sorted(list(set(factor_quadrant_map[key])))
        
        # Also get locations from the main dataset
        locations_raw = list(collection.distinct('Location'))
        def normalize_and_deduplicate(values):
            normalized_map = {}
            for v in values:
                if v:
                    v_stripped = str(v).strip()
                    if v_stripped:
                        v_normalized = v_stripped.lower()
                        if v_normalized not in normalized_map or len(v_stripped) < len(normalized_map[v_normalized]):
                            normalized_map[v_normalized] = v_stripped
            return sorted(set(normalized_map.values()))
        
        locations = normalize_and_deduplicate(locations_raw)
        
        return jsonify({
            'success': True,
            'data': {
                'quadrants': quadrants,
                'segments': segments,
                'factors': factors,
                'locations': locations,
                'mappings': {
                    'quadrant_segment': quadrant_segment_map,
                    'segment_factor': segment_factor_map,
                    'segment_quadrant': segment_quadrant_map,  # For auto-fill: segment -> quadrant
                    'factor_segment': factor_segment_map,  # For auto-fill: factor -> segment
                    'factor_quadrant': factor_quadrant_map  # For auto-fill: factor -> quadrant
                }
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

if __name__ == '__main__':
    # Create templates and static directories if they don't exist
    os.makedirs('templates', exist_ok=True)
    os.makedirs('static', exist_ok=True)
    
    print("Starting Flask application...")
    print("Open your browser and go to: http://127.0.0.1:5001")
    print("Press Ctrl+C to stop the server")
    
    app.run(debug=True, host='127.0.0.1', port=5001, use_reloader=False)
